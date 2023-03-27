import random
from Screen import initUI, draw_path_dijkstra, draw_path_a_star, canvas
from Dijkstra import dijkstra
from AStar import astar
import time as t
import openpyxl
from PIL import Image


file = r"C:\Users\jeroe\Desktop\Hundred.xlsx"

# def writexl(row,column, distance):
#     source_file = load_workbook(file)
#     sheet1 = source_file["Sheet1"]
#
#     sheet1.cell(row=row, column=column).value = distance
#     source_file.save(file)

# define main as a function. Have the point_num, con_low, con_high, double_chance and delta_fluctuation as random variabeles
# create grid edge distance for start- and end nodes

# settings
point_num = 12  # number of nodes - 2 <-> 10000
space_min = 50  # space between nodes in pixels
con_low = 2  # minimum number of connections 1 <-> 6 con_low has to be equal to or smaller than con_high
con_high = 4  # maximum number of connections - 2 <-> 10
double_chance = 40  # in percentage - 0 <-> 100
delta_fluctuation = 4 # linear random distance amplifier - 1 <-> 4
grid_edge_distance = 10 # the distance that the start and end node must have from the center of the grid

def save_as_png(canvas,file):
    # save postscipt image
    canvas.postscript(file = file + '.eps')
    # use PIL to convert to PNG
    img = Image.open(file + '.eps')
    img.save(file + '.png', 'png')

def exec(point_num,i,rota):
    graph = [[None for _ in range(point_num)] for _ in range(point_num)]
    points = []

    # launch environment
    initUI(points, point_num, space_min, con_low, con_high, double_chance, graph, delta_fluctuation, rota)

    # delta fluctuation

    # print(graph)

    # # print data
    # #
    # # print(points)
    # # print(delta)
    # # print(lines)
    # # print(graph)
    #
    #
    # Dijkstra's algorithm
    def call_dijkstra(start, end, graph):
        # dijkstra(graph, start, end)
        distance, path = dijkstra(graph, start, end)

        try:
            print("Dijkstra: Shortest Distance from %d to %d is %d" % (start, end, distance))
            print("Dijkstra: Shortest Path is: %s" % ' -> '.join(map(str, path)))
            return path

        except OverflowError:
            print("Dijkstra: No path")


    # A* algorithm
    def call_a_star(start, end, graph):
        # astar(graph, start, end)
        distance, path = astar(graph, start, end)
        if distance == 0:
            print("A*: No path")
            return distance
        else:
            print("A*: Shortest Distance from %d to %d is %d" % (start, end, distance))
            print("A*: Shortest Path is: %s" % ' -> '.join(map(str, path)))
            return path


    # call the algorithms
    start = 0
    end = 1

    tstart= t.time()
    # print(tstart) graph time

    # visualize the dijkstra path
    path_dijkstra = call_dijkstra(start, end, graph)
    if path_dijkstra != None:
        draw_path_dijkstra(path_dijkstra,points)
    # visualize the A* path
    # # dist = call_a_star(start, end, graph)
    #
    # path_a_star = call_a_star(start, end, graph)
    # if path_a_star != None:
    #     draw_path_a_star(path_a_star, points)

    time2 = t.time() - tstart
    row = i
    column = 4 + 4

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    def writenum():
        print('write to xlsx')
        chunk_size = 4000
        for i in range(0, len(graph), chunk_size):
            chunk = graph[i:i + chunk_size]
            for row in chunk:
                sheet.append(row)

        # save the workbook
        workbook.save(file)

    # writexl(row, column, time2)

    # if dist != 0:
    #     writenum()

    # print(graph)
        # visuals

    canvas.mainloop()  # show visuals

for i in range(1,2,1):
    print(f'Iteration = {i}')
    point_num = 100  # number of nodes - 2 <-> 10000
    rota = i
    exec(point_num,i,rota)

print('Done')
